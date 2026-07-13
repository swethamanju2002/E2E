from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from .models import Course,Placement, Service,CourseBooking, Contact,ClientProject,StudentReview
from .models import WorkshopPhoto, Certificate
from .models import Internship, ProcessStep
from .models import Internship
from .models import WorkshopRegistration
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.conf import settings
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone
from .models import UpcomingWorkshop

import os
def home(request):
    courses = Course.objects.all()
    services = Service.objects.all()
    projects=ClientProject.objects.all()
    reviews=StudentReview.objects.all()
    total_bookings = CourseBooking.objects.count()
    workshops       = WorkshopPhoto.objects.all()    
    certificates    = Certificate.objects.all()
    internships = Internship.objects.all()
    upcoming_workshop = UpcomingWorkshop.objects.first()
   
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        Contact.objects.create(name=name, email=email, message=message)
        return redirect('home')

    return render(request, 'index.html', {
        'courses': courses,
        'services': services,
        'projects':projects,
        'reviews':reviews,
        "total_bookings": total_bookings,
        'workshops': workshops,  
        'certificates':  certificates,
        'internships': internships,
        'upcoming_workshop': upcoming_workshop,
        

    })
    
from django.shortcuts import render, get_object_or_404
from .models import Course



def review_page(request):
    reviews = StudentReview.objects.all().order_by('-created_at')
    return render(request, 'reviews.html', {'reviews': reviews})
def course_detail(request, id):
    course = get_object_or_404(Course, id=id)

    description_lines = course.description.splitlines()

    return render(
        request,
        'courses.html',
        {
            'course': course,
            'description_lines': description_lines
        }
    )
def allcourse(request):
    allcourse = Course.objects.all()
    return render(request,'allcourse.html',{'allcourse':allcourse})

def about(request):
    return render(request, 'about.html')

def courses(request):
    courses = Course.objects.all()
    return render(request, 'courses.html', {'courses': courses})

def services(request):
    services = Service.objects.all()
    
    return render(request, 'services.html', {'services': services})

# views.py — replace service_detail
from .models import ProcessStep

def service_detail(request, pk):
    service = get_object_or_404(Service, pk=pk)
    return render(request, 'service_detail.html', {
        'service': service,
        'featured_demos': service.demo_links.filter(is_featured=True),
        'features': service.features.all(),
        'faqs': service.faqs.all(),
        'process_steps': ProcessStep.objects.all(),
    })
from .models import DemoCategory

def service_detail(request, pk):
    service = get_object_or_404(Service, pk=pk)
    demo_categories = DemoCategory.objects.filter(is_active=True)

    return render(request, 'service_detail.html', {
        'service': service,
        'demo_categories': demo_categories,
    })

def contact(request):
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        Contact.objects.create(name=name, email=email, message=message)
        return redirect('contact')
    return render(request, 'contact.html')


def placement(request):
    placements = Placement.objects.all()
    return render(request, 'placement.html', {
        'placements': placements
    })

from django.shortcuts import render
from .models import Career

def career_list(request):
    careers = Career.objects.all().order_by('-posted_on')
    return render(request, 'career.html', {'careers': careers})

def career_detail(request, pk):
    career = get_object_or_404(Career, pk=pk)
    return render(request, 'career_detail.html', {'career': career})


from django.shortcuts import render, get_object_or_404, redirect
from django.core.mail import send_mail
from django.conf import settings
from .models import Career
from .forms import JobApplicationForm
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from django.core.mail import EmailMessage


def apply_job(request, pk):
    career = get_object_or_404(Career, pk=pk)

    if request.method == 'POST':
        form = JobApplicationForm(request.POST, request.FILES)

        if form.is_valid():
            application = form.save(commit=False)
            application.career = career
            application.save()

            # ---------------- SAVE TO EXCEL ----------------
            job_folder = os.path.join(settings.MEDIA_ROOT, "job_applications")

            if not os.path.exists(job_folder):
                os.makedirs(job_folder)

            file_path = os.path.join(job_folder, "job_applications.xlsx")

            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append([
                    "Full Name",
                    "Email",
                    "Phone",
                    "Role Applied",
                    "Cover Letter",
                    "Resume File Name",
                    "Applied Date"
                ])

            sheet.append([
                application.full_name,
                application.email,
                application.mobile,
                career.role,
                application.cover_letter,
                application.resume.name,
                datetime.now().strftime('%d-%m-%Y %H:%M')
            ])

            workbook.save(file_path)

            # ---------------- EMAIL TO CANDIDATE ----------------
            send_mail(
                subject="Application Received - Errors2Experts",
                message=f"""
Hi {application.full_name},

Greetings from Errors2Experts!

We have received your application for the role of {career.role}.
Our HR team will review your profile and contact you soon.

Thank you for your interest.

Regards,
Errors2Experts Team
""",
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[application.email],
            )

            # ---------------- EMAIL TO ADMIN (YOU) ----------------
            admin_subject = f"🚀 New Job Application - {career.role}"

            admin_message = f"""
New Job Application Received!

Candidate Details:

Name: {application.full_name}
Email: {application.email}
Phone: {application.mobile}

Role Applied: {career.role}

Cover Letter:
{application.cover_letter}

Applied On: {datetime.now().strftime('%d-%m-%Y %H:%M')}

Please review the attached resume.

Errors2Experts System
"""

            admin_email = EmailMessage(
                subject=admin_subject,
                body=admin_message,
                from_email=settings.EMAIL_HOST_USER,
                to=[settings.ADMIN_NOTIFICATION_EMAIL],
            )

            # Attach Resume File
            if application.resume:
                admin_email.attach_file(application.resume.path)

            admin_email.send()

            return render(request, 'application_success.html')

    else:
        form = JobApplicationForm()

    return render(request, 'apply_job.html', {'form': form, 'career': career})
import os
from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import JsonResponse


def demo_booking(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        education = request.POST.get("education")
        source = request.POST.get("source")

        # ===============================
        # ✅ 1️⃣ STORE DATA IN EXCEL
        # ===============================

        file_path = os.path.join(settings.BASE_DIR, "demo_booking1.xlsx")

        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Demo Bookings"

            ws.append([
                "Name",
                "Email",
                "Mobile",
                "Education",
                "Source"
            ])
            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        ws.append([
            name,
            email,
            mobile,
            education,
            source
        ])

        wb.save(file_path)

        # ===============================
        # ✅ 2️⃣ SEND EMAIL TO ADMIN
        # ===============================

        admin_subject = "New Demo Booking - Errors2Experts"

        admin_message = f"""
New Demo Booking Received

Name: {name}
Email: {email}
Mobile: {mobile}
Education: {education}
Source: {source}
"""

        EmailMessage(
            admin_subject,
            admin_message,
            settings.EMAIL_HOST_USER,
            [settings.EMAIL_HOST_USER],
        ).send()

        # ===============================
        # ✅ 3️⃣ SEND WELCOME EMAIL TO USER
        # ===============================

        user_subject = "Welcome to Errors2Experts 🚀"

        user_message = f"""
Hi {name},

Greetings from Errors2Experts!

Thank you for showing interest in learning with us.
We have received your demo booking successfully.

Our team will contact you soon and guide you through the next steps.

At Errors2Experts, we believe:
"Learn from errors, grow with knowledge, and become an expert."

Stay connected with us:

🌐 Website: https://yourwebsite.com
📸 Instagram: https://www.instagram.com/errors2experts_2026/

Let’s learn, grow, and build your future together!

Best Regards,
Team Errors2Experts
"""

        EmailMessage(
            user_subject,
            user_message,
            settings.EMAIL_HOST_USER,
            [email],
        ).send()

        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "failed"})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Course, CourseBooking
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime, timedelta

def course_detail(request, id):
    course = get_object_or_404(Course, id=id)
    return render(request, "courses.html", {"course": course})


from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.core.mail import EmailMessage
from .models import Course, CourseBooking
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os


def book_course(request, id):
    course = get_object_or_404(Course, id=id)

    if request.method == "POST":

        name = request.POST.get("name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        education = request.POST.get("education")
        year_passed = request.POST.get("year_passed")
        payment_type = "full"

        price = int(course.price)

        # ---------------- PAYMENT CALCULATION ----------------
        if payment_type == "emi":
            amount = price // 2
            balance = price - amount
            due_date = datetime.now() + timedelta(days=15)
        else:
            amount = price
            balance = 0
            due_date = None

        # ---------------- SAVE TO DATABASE ----------------
        CourseBooking.objects.create(
            name=name,
            email=email,
            mobile=mobile,
            education=education,
            year_passed=year_passed,
            course=course.title,
            amount=amount,
            
        )

        # ---------------- SAVE TO EXCEL ----------------
        booking_folder = os.path.join(settings.BASE_DIR, "media", "bookings")
        if not os.path.exists(booking_folder):
         os.makedirs(booking_folder)

        file_path = os.path.join(booking_folder, "course_bookings.xlsx")

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "Name",
                "Email",
                "Mobile",
                "Education",
                "Year Passed",
                "Course",
                "Payment Type",
                "First Payment",
                "Balance",
                "Due Date",
                "Booking Date"
            ])

        sheet.append([
            name,
            email,
            mobile,
            education,
            year_passed,
            course.title,
            payment_type,
            amount,
            balance,
            due_date.strftime('%d-%m-%Y') if due_date else "N/A",
            datetime.now().strftime('%d-%m-%Y %H:%M')
        ])

        workbook.save(file_path)

        # ---------------- STUDENT EMAIL CONTENT ----------------
        if payment_type == "emi":
            payment_message = f"""
Your First Payment: ₹{amount}
Remaining Balance: ₹{balance}
Next Payment Due Date: {due_date.strftime('%d-%m-%Y')}

Kindly complete your first payment and send the screenshot via:

WhatsApp: +919363342646
Email: errors2experts.official@gmail.com


Invoice is mandatory.
"""
        else:
            payment_message = f"""
Total Amount: ₹{amount}

Kindly complete your payment and send the screenshot via:

WhatsApp: +919363342646
Email: errors2experts.official@gmail.com


Invoice is mandatory.
"""

        student_message = f"""
Hi {name},

Welcome to Errors2Experts!

Course: {course.title}

{payment_message}

Terms & Conditions:

1. Errors2Experts is a MSME registered training institute.
2. Amount once paid is non-refundable.
3. Maintain discipline and behave professionally.
4. If taking leave, inform in course WhatsApp group.
5. More than 2 days leave requires proof + email.
6. Placement assistance provided (not guarantee).

Let's Learn & Grow Together 🚀

Thank you,
Errors2Experts Team
"""

        # ---------------- SEND EMAIL TO STUDENT ----------------
        student_email = EmailMessage(
            subject="Course Booking Confirmation - Errors2Experts",
            body=student_message,
            from_email=settings.EMAIL_HOST_USER,
            to=[email],
        )

        # Attach QR
        qr_path = os.path.join(settings.BASE_DIR, "static/images/qr.jpeg")
        if os.path.exists(qr_path):
            student_email.attach_file(qr_path)

        student_email.send()

        # ---------------- SEND EMAIL TO ADMIN ----------------
        admin_subject = "🚀 New Course Booking Received"

        admin_message = f"""
New Booking Alert!

Student Details:

Name: {name}
Email: {email}
Mobile: {mobile}
Education: {education}
Year Passed: {year_passed}

Course: {course.title}
Payment Type: {payment_type}
First Payment: ₹{amount}
Balance: ₹{balance}
Due Date: {due_date.strftime('%d-%m-%Y') if due_date else 'N/A'}

Booking Time: {datetime.now().strftime('%d-%m-%Y %H:%M')}

Please follow up with the student.

Errors2Experts System
"""

        admin_email = EmailMessage(
            subject=admin_subject,
            body=admin_message,
            from_email=settings.EMAIL_HOST_USER,
            to=[settings.ADMIN_NOTIFICATION_EMAIL],  # add in settings.py
        )

        admin_email.send()

        # ---------------- SHOW QR PAGE ----------------
        return render(request, "show_qr.html", {
            "amount": amount,
            "balance": balance,
            "due_date": due_date,
            "payment_type": payment_type,
            "course_title": course.title,
        })

    return redirect("course_detail", id=id)

import os
from django.conf import settings
from django.shortcuts import redirect
from django.core.mail import send_mail
from openpyxl import Workbook, load_workbook
from .models import ServiceBooking
from django.contrib import messages


def service_booking(request):
    if request.method == "POST":

        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        service_name = request.POST.get("service_name")
        message = request.POST.get("message")
        preferred_date = request.POST.get("preferred_date")

        # Save to database
        booking = ServiceBooking.objects.create(
            full_name=full_name,
            email=email,
            mobile=mobile,
            service_name=service_name,
            message=message,
            preferred_date=preferred_date
        )

        # Save to Excel
        file_path = os.path.join(settings.BASE_DIR, "service_bookings.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Full Name", "Email", "Mobile", "Service", "Preferred Date", "Message"])

        ws.append([full_name, email, mobile, service_name, preferred_date, message])
        wb.save(file_path)

        # Email to Admin
        admin_subject = "New Service Booking"
        admin_message = f"""
New Service Booking Received

Name: {full_name}
Email: {email}
Mobile: {mobile}
Service: {service_name}
Preferred Date: {preferred_date}
Message: {message}
"""

        send_mail(
            admin_subject,
            admin_message,
            settings.EMAIL_HOST_USER,
            [settings.EMAIL_HOST_USER],
            fail_silently=False,
        )

        # Email to Client
        client_subject = "Thank You for Registering"
        client_message = f"""
Hi {full_name},

Thank you for registering for our {service_name} service.

Our team will contact you as soon as possible.

Stay connected with us:

🌐 Website: https://yourwebsite.com
📸 Instagram: https://www.instagram.com/errors2experts_2026/

Let’s learn, grow, and build your future together!

Best Regards,
Errors2Experts Team
"""

        send_mail(
            client_subject,
            client_message,
            settings.EMAIL_HOST_USER,
            [email],
            fail_silently=False,
        )
        messages.success(request, "Service Registered Successfully")
        return redirect("services")  # change if needed
    
from .models import WorkshopPhoto, UpcomingWorkshop

def workshop_gallery(request):
    workshops = WorkshopPhoto.objects.all()

    upcoming_workshop = UpcomingWorkshop.objects.order_by("event_date").first()

    return render(request, "workshop.html", {
        "workshops": workshops,
        "upcoming_workshop": upcoming_workshop,
    })
 
def certificate_gallery(request):
    """Dedicated full-page certificate gallery."""
    certificates = Certificate.objects.all()
    return render(request, 'certificate.html', {'certificates': certificates})
from django.shortcuts import render, get_object_or_404
from collections import defaultdict
from .models import Internship

# View for the landing page (Cards)
def internship_list(request):
    internships = Internship.objects.all()
    return render(request, 'all_internship.html', {'internships': internships})

# View for the roadmap detail page
def internship_detail(request, pk):
    internship = get_object_or_404(Internship, pk=pk)
    
    # Logic to group days into weekly timeline blocks
    timeline_data = defaultdict(list)
    for index, topic in enumerate(internship.syllabus):
        week = (index // 5) + 1  
        timeline_data[week].append({'day': index + 1, 'topic': topic})
        
    return render(request, 'internship_detail.html', {
        'internship': internship, 
        'timeline_data': dict(timeline_data)
    })
    
# from django.core.paginator import Paginator
# from django.db.models import Q
# from .models import ServiceDemoLink

# def live_demo(request):
#     demos = (ServiceDemoLink.objects
#              .select_related('service')
#              .exclude(url__isnull=True).exclude(url='')
#              .order_by('order', 'service__title'))

#     query = request.GET.get('q', '').strip()
#     category = request.GET.get('category', 'all').strip()

#     if query:
#         demos = demos.filter(
#             Q(title__icontains=query) |
#             Q(description__icontains=query) |
#             Q(technologies__icontains=query)
#         )
#     if category and category.lower() != 'all':
#         demos = demos.filter(category__iexact=category)

#     categories = (ServiceDemoLink.objects
#                   .exclude(category='')
#                   .values_list('category', flat=True)
#                   .distinct().order_by('category'))

#     paginator = Paginator(demos, 9)
#     page_obj = paginator.get_page(request.GET.get('page'))

#     context = {
#         'page_obj': page_obj,
#         'categories': categories,
#         'query': query,
#         'active_category': category or 'all',
#     }

#     # AJAX request → return only the results fragment, no full-page render
#     if request.headers.get('x-requested-with') == 'XMLHttpRequest':
#         return render(request, 'partials/live_demo_results.html', context)

#     return render(request, 'live_demo.html', context)
 

def workshop_registration(request):

    if request.method == "POST":

        name = request.POST.get("full_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        education = request.POST.get("education")
        workshop = request.POST.get("workshop")

        # ================= DATABASE =================

        WorkshopRegistration.objects.create(
            full_name=name,
            email=email,
            mobile=mobile,
            education=education,
            workshop=workshop
        )

        # ================= EXCEL =================

        file_path = os.path.join(
            settings.BASE_DIR,
            "workshop_registrations.xlsx"
        )

        if os.path.exists(file_path):

            wb = load_workbook(file_path)
            ws = wb.active

        else:

            wb = Workbook()
            ws = wb.active

            ws.append([
                "Name",
                "Email",
                "Mobile",
                "Education",
                "Workshop",
                "Date"
            ])

        ws.append([
            name,
            email,
            mobile,
            education,
            workshop,
            datetime.now().strftime("%d-%m-%Y %H:%M")
        ])

        wb.save(file_path)

        # ================= ADMIN MAIL =================

        admin_subject = f"New Workshop Registration - {workshop}"

        admin_message = f"""
New Workshop Registration

Name      : {name}
Email     : {email}
Mobile    : {mobile}
Education : {education}
Workshop  : {workshop}

Registration Time :
{datetime.now().strftime("%d-%m-%Y %H:%M")}
"""

        EmailMessage(
            subject=admin_subject,
            body=admin_message,
            from_email=settings.EMAIL_HOST_USER,
            to=[settings.ADMIN_NOTIFICATION_EMAIL]
        ).send()

        # ================= USER MAIL =================

        subject = f"Workshop Registration Successful - {workshop}"

        text_content = f"""
Hi {name},

Thank you for registering successfully for the {workshop}.

Your registration has been received successfully.

Our team will contact you shortly with the meeting link and workshop details.

Regards,
Errors2Experts Team
"""

        html_content = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
</head>

<body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:30px;">

<div style="max-width:650px;margin:auto;background:#ffffff;border-radius:10px;padding:30px;border:1px solid #ddd;">

<h2 style="color:#2e7d32;">
Hello {name},
</h2>

<p>
Thank you for registering for our
<b>{workshop}</b>.
</p>

<p>
Your registration has been received successfully.
</p>

<p>
Our team will contact you shortly with the workshop meeting link and complete instructions.
</p>

<hr>

<h3 style="color:#1b9615;">
Workshop Details
</h3>

<p>

<b>Workshop :</b> {workshop}<br>
<b>Status :</b> Registration Confirmed ..

</p>

<br>

<p>
Thank you for choosing
<b>Errors2Experts</b>.
</p>

<p>
Regards,<br>
<b>Errors2Experts Team</b>
</p>

</div>

</body>
</html>
"""

        mail = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=settings.EMAIL_HOST_USER,
            to=[email],
        )

        mail.attach_alternative(html_content, "text/html")
        mail.send()

        # ================= REDIRECT =================

        return redirect("home")

    return redirect("home")

from .models import DemoRequest, DemoCategory
from django.core.mail import EmailMessage
from django.contrib import messages


def demo_request(request):
    if request.method == "POST":

        organization_name = request.POST.get("organization_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        category_value = request.POST.get("category")
        custom_requirement = request.POST.get("custom_requirement")

        category = None

        if category_value == "other":
            category = None
        else:
            try:
                category = DemoCategory.objects.get(id=category_value)
            except DemoCategory.DoesNotExist:
                category = None

        # Save Request
        DemoRequest.objects.create(
            organization_name=organization_name,
            email=email,
            mobile=mobile,
            category=category,
            custom_requirement=custom_requirement,
        )

        
        # CATEGORY FOUND
        

        if category:

            user_subject = f"{category.name if category else 'Other'} Demo Link"

            user_message = f"""
Hi {organization_name},

Thank you for requesting the {category.name if category else "Other"} demo.

Below is your demo link:

{category.demo_link}

Thank you.

Regards,
Errors2Experts Team
"""

            EmailMessage(
                user_subject,
                user_message,
                settings.EMAIL_HOST_USER,
                [email],
            ).send()

            messages.success(
                request,
                "Demo link has been sent to your email."
            )

        
        # OTHER REQUIREMENT
        

        else:

            # User Mail

            EmailMessage(
                "Demo Request Received",
                f"""
Hi {organization_name},

Thank you for contacting Errors2Experts.

We have received your custom demo request.

Our technical team will review your requirement.

You will receive a response within 2-3 business days.

Regards,
Errors2Experts Team
""",
                settings.EMAIL_HOST_USER,
                [email],
            ).send()

            # Admin Mail

            EmailMessage(
                "🚨 Critical Custom Demo Request",
                f"""
Customer Name : {organization_name}

Email : {email}

Mobile : {mobile}

Requirement :

{custom_requirement}
""",
                settings.EMAIL_HOST_USER,
                [settings.ADMIN_NOTIFICATION_EMAIL],
            ).send()

            messages.success(
                request,
                "Your request has been received successfully."
            )

        return redirect("services")

    return redirect("services")
